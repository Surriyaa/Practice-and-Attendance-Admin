import openpyxl

# Setup logger
from com.bridgelabz.utilities.logger import Logger
logger = Logger.get_logger("QuestionLoader")

def load_questions_from_excel(file_path):
    logger.info(f"Loading Excel file: {file_path}")
    workbook = openpyxl.load_workbook(file_path)
    data = {}

    for sheet in workbook.sheetnames:
        logger.info(f"Reading sheet: {sheet}")
        worksheet = workbook[sheet]
        questions = []
        for row in worksheet.iter_rows(min_row=2, max_col=1, values_only=True):
            question = row[0]
            if question:
                logger.info(f"Found question in {sheet}: {question}")
                questions.append(question)
        data[sheet.lower()] = questions  # e.g., "Level1" → "level1"

    logger.info("Finished loading all levels")
    return data


def read_practice_head_data(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        name, email, mobile, base_coe, expected_result = row
        data.append((name, email, str(mobile), base_coe, expected_result))
    return data
